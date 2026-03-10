"""Report writer for Excel and CSV output."""
import csv
from datetime import datetime
from pathlib import Path

import structlog
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.exceptions import ReportGenerationError
from src.domain.models import (
    ComparisonResult,
    ComparisonSummary,
    DuplicateRecord,
    MatchStatus,
    ProductCount,
)

logger = structlog.get_logger()


class ReportWriter:
    """Report writer for generating comparison reports."""

    def __init__(self, output_dir: Path | str) -> None:
        """
        Initialize report writer.

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _generate_filename(self, extension: str) -> Path:
        """Generate filename with timestamp."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.output_dir / f"comparison_report_{timestamp}.{extension}"

    def _style_header(self, ws: Worksheet, row: int = 1) -> None:
        """Apply header styling to worksheet."""
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """Auto-adjust column widths."""
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                if isinstance(cell, Cell) and cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                    except (TypeError, AttributeError):
                        pass
            adjusted_width = min(max_length + 2, 50)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width

    def write_excel_report(
        self,
        summary: ComparisonSummary,
        product_counts: list[ProductCount],
        comparison_results: list[ComparisonResult],
        duplicates: list[DuplicateRecord],
    ) -> Path:
        """
        Write Excel report with multiple sheets.

        Args:
            summary: Comparison summary
            product_counts: Product count aggregation
            comparison_results: Individual comparison results
            duplicates: Duplicate record information

        Returns:
            Path to generated Excel file
        """
        try:
            wb = Workbook()

            # Sheet 1: Summary
            ws_summary = wb.active
            if ws_summary is None:
                ws_summary = wb.create_sheet("Summary")
            ws_summary.title = "Summary"
            self._write_summary_sheet(ws_summary, summary)

            # Sheet 2: Product Counts
            ws_products = wb.create_sheet("Product Counts")
            self._write_product_counts_sheet(ws_products, product_counts)

            # Sheet 3: Matched
            ws_matched = wb.create_sheet("Matched")
            matched = [
                r for r in comparison_results if r.status == MatchStatus.MATCHED
            ]
            self._write_matched_sheet(ws_matched, matched)

            # Sheet 4: Name Mismatches
            ws_mismatch = wb.create_sheet("Name Mismatches")
            mismatches = [
                r for r in comparison_results if r.status == MatchStatus.NAME_MISMATCH
            ]
            self._write_mismatches_sheet(ws_mismatch, mismatches)

            # Sheet 5: Source Only
            ws_source_only = wb.create_sheet("Source Only")
            source_only = [
                r for r in comparison_results if r.status == MatchStatus.SOURCE_ONLY
            ]
            self._write_single_source_sheet(
                ws_source_only, source_only, "Source Product"
            )

            # Sheet 6: Target Only
            ws_target_only = wb.create_sheet("Target Only")
            target_only = [
                r for r in comparison_results if r.status == MatchStatus.TARGET_ONLY
            ]
            self._write_single_source_sheet(
                ws_target_only, target_only, "Target Product", is_target=True
            )

            # Sheet 7: Duplicates
            ws_duplicates = wb.create_sheet("Duplicates")
            self._write_duplicates_sheet(ws_duplicates, duplicates)

            # Save workbook
            file_path = self._generate_filename("xlsx")
            wb.save(file_path)

            logger.info("excel_report_generated", path=str(file_path))
            return file_path

        except Exception as e:
            raise ReportGenerationError(f"Failed to generate Excel report: {e}") from e

    def _write_summary_sheet(
        self, ws: Worksheet, summary: ComparisonSummary
    ) -> None:
        """Write summary sheet."""
        ws.append(["Metric", "Value"])
        self._style_header(ws)

        # Data metrics
        metrics = [
            ("Total Source Records", summary.total_source),
            ("Total Target Records", summary.total_target),
            ("", ""),
            ("Extra Duplicate Source", summary.extra_duplicate_source_count),
            ("Extra Duplicate Target", summary.extra_duplicate_target_count),
            ("", ""),
            ("Unique Source Serials", summary.unique_source_count),
            ("Unique Target Serials", summary.unique_target_count),
            ("", ""),
            ("Matched", summary.matched_count),
            ("Name Mismatch", summary.name_mismatch_count),
            ("Source Only", summary.source_only_count),
            ("Target Only", summary.target_only_count),
            ("", ""),
            ("Validation", ""),
            (
                f"Source: {summary.total_source} = {summary.unique_source_count} + {summary.extra_duplicate_source_count}",
                "PASS" if summary.validate_consistency() else "FAIL",
            ),
            (
                f"Target: {summary.total_target} = {summary.unique_target_count} + {summary.extra_duplicate_target_count}",
                "PASS" if summary.validate_consistency() else "FAIL",
            ),
        ]

        for metric, value in metrics:
            ws.append([metric, value])

        self._auto_adjust_columns(ws)

    def _write_product_counts_sheet(
        self, ws: Worksheet, product_counts: list[ProductCount]
    ) -> None:
        """Write product counts sheet."""
        ws.append(["Product Name", "Source Count", "Target Count", "Difference"])
        self._style_header(ws)

        for pc in product_counts:
            ws.append([pc.product_name, pc.source_count, pc.target_count, pc.diff])

        self._auto_adjust_columns(ws)

    def _write_matched_sheet(
        self, ws: Worksheet, matched: list[ComparisonResult]
    ) -> None:
        """Write matched records sheet."""
        ws.append(["Serial Number", "Product Name"])
        self._style_header(ws)

        for r in matched:
            ws.append([r.serial_number, r.source_product_name])

        self._auto_adjust_columns(ws)

    def _write_mismatches_sheet(
        self, ws: Worksheet, mismatches: list[ComparisonResult]
    ) -> None:
        """Write name mismatches sheet."""
        ws.append(["Serial Number", "Source Product", "Target Product"])
        self._style_header(ws)

        for r in mismatches:
            ws.append([r.serial_number, r.source_product_name, r.target_product_name])

        self._auto_adjust_columns(ws)

    def _write_single_source_sheet(
        self,
        ws: Worksheet,
        results: list[ComparisonResult],
        product_column: str,
        is_target: bool = False,
    ) -> None:
        """Write source-only or target-only sheet."""
        ws.append(["Serial Number", product_column])
        self._style_header(ws)

        for r in results:
            product = r.target_product_name if is_target else r.source_product_name
            ws.append([r.serial_number, product])

        self._auto_adjust_columns(ws)

    def _write_duplicates_sheet(
        self, ws: Worksheet, duplicates: list[DuplicateRecord]
    ) -> None:
        """Write duplicates sheet."""
        ws.append(["Serial Number", "Source", "Occurrence Count", "Product Names"])
        self._style_header(ws)

        for d in duplicates:
            ws.append([
                d.serial_number,
                d.source,
                d.occurrence_count,
                ", ".join(d.product_names),
            ])

        self._auto_adjust_columns(ws)

    def write_csv_reports(
        self,
        summary: ComparisonSummary,
        product_counts: list[ProductCount],
        comparison_results: list[ComparisonResult],
        duplicates: list[DuplicateRecord],
    ) -> list[Path]:
        """
        Write CSV reports (multiple files).

        Returns:
            List of paths to generated CSV files
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_files: list[Path] = []

            # Summary CSV
            summary_path = self.output_dir / f"summary_{timestamp}.csv"
            with open(summary_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Metric", "Value"])
                writer.writerow(["Total Source Records", summary.total_source])
                writer.writerow(["Total Target Records", summary.total_target])
                writer.writerow(["Extra Duplicate Source", summary.extra_duplicate_source_count])
                writer.writerow(["Extra Duplicate Target", summary.extra_duplicate_target_count])
                writer.writerow(["Unique Source Serials", summary.unique_source_count])
                writer.writerow(["Unique Target Serials", summary.unique_target_count])
                writer.writerow(["Matched", summary.matched_count])
                writer.writerow(["Name Mismatch", summary.name_mismatch_count])
                writer.writerow(["Source Only", summary.source_only_count])
                writer.writerow(["Target Only", summary.target_only_count])
            output_files.append(summary_path)

            # Product counts CSV
            products_path = self.output_dir / f"product_counts_{timestamp}.csv"
            with open(products_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Product Name", "Source Count", "Target Count", "Difference"])
                for pc in product_counts:
                    writer.writerow([pc.product_name, pc.source_count, pc.target_count, pc.diff])
            output_files.append(products_path)

            # Name mismatches CSV
            mismatches = [
                r for r in comparison_results if r.status == MatchStatus.NAME_MISMATCH
            ]
            mismatches_path = self.output_dir / f"name_mismatches_{timestamp}.csv"
            with open(mismatches_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Serial Number", "Source Product", "Target Product"])
                for r in mismatches:
                    writer.writerow([r.serial_number, r.source_product_name, r.target_product_name])
            output_files.append(mismatches_path)

            # Duplicates CSV
            duplicates_path = self.output_dir / f"duplicates_{timestamp}.csv"
            with open(duplicates_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Serial Number", "Source", "Occurrence Count", "Product Names"])
                for d in duplicates:
                    writer.writerow([d.serial_number, d.source, d.occurrence_count, "|".join(d.product_names)])
            output_files.append(duplicates_path)

            logger.info("csv_reports_generated", count=len(output_files))
            return output_files

        except Exception as e:
            raise ReportGenerationError(f"Failed to generate CSV reports: {e}") from e

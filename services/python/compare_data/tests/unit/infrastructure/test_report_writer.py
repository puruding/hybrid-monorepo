"""Unit tests for report writer."""
from pathlib import Path

import pytest
from openpyxl import load_workbook
from src.domain.models import (
    ComparisonResult,
    ComparisonSummary,
    DuplicateRecord,
    MatchStatus,
    ProductCount,
)
from src.infrastructure.report_writer import ReportWriter


@pytest.fixture
def sample_summary() -> ComparisonSummary:
    """Sample comparison summary."""
    return ComparisonSummary(
        total_source=100,
        total_target=95,
        extra_duplicate_source_count=2,
        extra_duplicate_target_count=1,
        unique_source_count=98,
        unique_target_count=94,
        matched_count=88,
        name_mismatch_count=5,
        source_only_count=5,
        target_only_count=1,
    )


@pytest.fixture
def sample_product_counts() -> list[ProductCount]:
    """Sample product counts."""
    return [
        ProductCount("Cisco ASA", 30, 28, 2),
        ProductCount("FortiGate", 50, 52, -2),
        ProductCount("Palo Alto", 20, 14, 6),
    ]


@pytest.fixture
def sample_comparison_results() -> list[ComparisonResult]:
    """Sample comparison results."""
    return [
        ComparisonResult("SN001", MatchStatus.MATCHED, "Product A", "Product A"),
        ComparisonResult("SN002", MatchStatus.NAME_MISMATCH, "Product B", "Product X"),
        ComparisonResult("SN003", MatchStatus.SOURCE_ONLY, "Product C", None),
        ComparisonResult("SN004", MatchStatus.TARGET_ONLY, None, "Product D"),
    ]


@pytest.fixture
def sample_duplicates() -> list[DuplicateRecord]:
    """Sample duplicate records."""
    return [
        DuplicateRecord("SN100", ["Product A", "Product B"], 2, "source"),
        DuplicateRecord("SN200", ["Product X", "Product Y", "Product Z"], 3, "target"),
    ]


class TestReportWriter:
    """Test cases for ReportWriter."""

    def test_write_excel_report(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
        sample_product_counts: list[ProductCount],
        sample_comparison_results: list[ComparisonResult],
        sample_duplicates: list[DuplicateRecord],
    ) -> None:
        """Test Excel report generation."""
        writer = ReportWriter(output_dir)

        file_path = writer.write_excel_report(
            sample_summary,
            sample_product_counts,
            sample_comparison_results,
            sample_duplicates,
        )

        # Verify file exists
        assert file_path.exists()
        assert file_path.suffix == ".xlsx"

        # Verify workbook structure
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames

        assert "Summary" in sheet_names
        assert "Product Counts" in sheet_names
        assert "Name Mismatches" in sheet_names
        assert "Source Only" in sheet_names
        assert "Target Only" in sheet_names
        assert "Duplicates" in sheet_names

    def test_write_excel_report_summary_content(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
        sample_product_counts: list[ProductCount],
        sample_comparison_results: list[ComparisonResult],
        sample_duplicates: list[DuplicateRecord],
    ) -> None:
        """Test Excel summary sheet content."""
        writer = ReportWriter(output_dir)

        file_path = writer.write_excel_report(
            sample_summary,
            sample_product_counts,
            sample_comparison_results,
            sample_duplicates,
        )

        wb = load_workbook(file_path)
        ws = wb["Summary"]

        # Check header
        assert ws.cell(1, 1).value == "Metric"
        assert ws.cell(1, 2).value == "Value"

        # Check some data values
        # Find the row with "Total Source Records"
        found_source = False
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[0].value == "Total Source Records":
                assert row[1].value == 100
                found_source = True
                break
        assert found_source

    def test_write_excel_report_product_counts_content(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
        sample_product_counts: list[ProductCount],
        sample_comparison_results: list[ComparisonResult],
        sample_duplicates: list[DuplicateRecord],
    ) -> None:
        """Test Excel product counts sheet content."""
        writer = ReportWriter(output_dir)

        file_path = writer.write_excel_report(
            sample_summary,
            sample_product_counts,
            sample_comparison_results,
            sample_duplicates,
        )

        wb = load_workbook(file_path)
        ws = wb["Product Counts"]

        # Check header
        assert ws.cell(1, 1).value == "Product Name"
        assert ws.cell(1, 2).value == "Source Count"
        assert ws.cell(1, 3).value == "Target Count"
        assert ws.cell(1, 4).value == "Difference"

        # Check data (row 2 = first data row)
        assert ws.cell(2, 1).value == "Cisco ASA"
        assert ws.cell(2, 2).value == 30
        assert ws.cell(2, 3).value == 28
        assert ws.cell(2, 4).value == 2

    def test_write_excel_report_name_mismatches_content(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
        sample_product_counts: list[ProductCount],
        sample_comparison_results: list[ComparisonResult],
        sample_duplicates: list[DuplicateRecord],
    ) -> None:
        """Test Excel name mismatches sheet content."""
        writer = ReportWriter(output_dir)

        file_path = writer.write_excel_report(
            sample_summary,
            sample_product_counts,
            sample_comparison_results,
            sample_duplicates,
        )

        wb = load_workbook(file_path)
        ws = wb["Name Mismatches"]

        # Check header
        assert ws.cell(1, 1).value == "Serial Number"
        assert ws.cell(1, 2).value == "Source Product"
        assert ws.cell(1, 3).value == "Target Product"

        # Check data - only 1 mismatch in sample
        assert ws.cell(2, 1).value == "SN002"
        assert ws.cell(2, 2).value == "Product B"
        assert ws.cell(2, 3).value == "Product X"

    def test_write_csv_reports(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
        sample_product_counts: list[ProductCount],
        sample_comparison_results: list[ComparisonResult],
        sample_duplicates: list[DuplicateRecord],
    ) -> None:
        """Test CSV report generation."""
        writer = ReportWriter(output_dir)

        file_paths = writer.write_csv_reports(
            sample_summary,
            sample_product_counts,
            sample_comparison_results,
            sample_duplicates,
        )

        # Should generate 4 CSV files
        assert len(file_paths) == 4

        # Verify all files exist
        for path in file_paths:
            assert path.exists()
            assert path.suffix == ".csv"

        # Verify file names contain expected prefixes
        file_names = [p.name for p in file_paths]
        assert any("summary" in name for name in file_names)
        assert any("product_counts" in name for name in file_names)
        assert any("name_mismatches" in name for name in file_names)
        assert any("duplicates" in name for name in file_names)

    def test_output_directory_created(self, tmp_path: Path) -> None:
        """Test that output directory is created if it doesn't exist."""
        new_dir = tmp_path / "new_output_dir"
        assert not new_dir.exists()

        writer = ReportWriter(new_dir)

        assert new_dir.exists()

    def test_empty_results(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
    ) -> None:
        """Test report generation with empty results."""
        writer = ReportWriter(output_dir)

        # Empty lists
        empty_products: list[ProductCount] = []
        empty_results: list[ComparisonResult] = []
        empty_duplicates: list[DuplicateRecord] = []

        file_path = writer.write_excel_report(
            sample_summary,
            empty_products,
            empty_results,
            empty_duplicates,
        )

        assert file_path.exists()

        wb = load_workbook(file_path)
        # Product Counts sheet should only have header
        ws_products = wb["Product Counts"]
        assert ws_products.max_row == 1  # Only header

    def test_duplicates_sheet_content(
        self,
        output_dir: Path,
        sample_summary: ComparisonSummary,
        sample_product_counts: list[ProductCount],
        sample_comparison_results: list[ComparisonResult],
        sample_duplicates: list[DuplicateRecord],
    ) -> None:
        """Test Excel duplicates sheet content."""
        writer = ReportWriter(output_dir)

        file_path = writer.write_excel_report(
            sample_summary,
            sample_product_counts,
            sample_comparison_results,
            sample_duplicates,
        )

        wb = load_workbook(file_path)
        ws = wb["Duplicates"]

        # Check header
        assert ws.cell(1, 1).value == "Serial Number"
        assert ws.cell(1, 2).value == "Source"
        assert ws.cell(1, 3).value == "Occurrence Count"
        assert ws.cell(1, 4).value == "Product Names"

        # Check first duplicate data
        assert ws.cell(2, 1).value == "SN100"
        assert ws.cell(2, 2).value == "source"
        assert ws.cell(2, 3).value == 2
        assert ws.cell(2, 4).value == "Product A, Product B"
